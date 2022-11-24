from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
import sys
import os
from investment_generator import edit_spreadsheet


class MainWindow(ttk.Frame):
    def new_spreadsheet_win(self):
        NewSpreadsheet(self.parent)

    def existing_spreadsheet_win(self):
        ExistingSpreadsheet(self.parent)

    def __init__(self, parent):
        ttk.Frame.__init__(self, parent)
        self.parent = parent
        self.parent.columnconfigure(0, weight=1)
        self.parent.rowconfigure(0, weight=1)
        self.parent.title('Investment Spreadsheet Generator')

        self.grid(row=0, column=0, sticky=(N, S, E, W))

        screen_size = self.parent.wm_maxsize()
        root_size = (screen_size[0]/2, screen_size[1]/2, screen_size[0]/4, screen_size[1]/14)
        root_size = [int(each) for each in root_size]
        self.parent.geometry(f'{root_size[0]}x{root_size[1]}+{root_size[2]}+{root_size[3]}')

        self.header = ttk.Label(self, text='Welcome to the Investment Spreadsheet Generator', font='TkCaptionFont')
        self.header.grid(row=0, column=0, columnspan=2)
        self.header.configure(padding='15')

        self.new_spreadsheet = ttk.Button(self, text='Create a New Spreadsheet', command=self.new_spreadsheet_win)
        self.new_spreadsheet.grid(row=1, column=0, padx='10', pady='10')

        self.existing_spreadsheet = ttk.Button(self, text='Edit an Existing Spreadsheet',
                                               command=self.existing_spreadsheet_win)

        self.existing_spreadsheet.grid(row=1, column=1)
        self.existing_spreadsheet.grid(row=1, column=1, padx='10', pady='10')

        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)


class NewSpreadsheet(Toplevel):
    def __init__(self, parent):
        Toplevel.__init__(self, parent)
        self.parent = parent
        self.parent.withdraw()
        self.title('New Investment Spreadsheet')

        screen_size = self.wm_maxsize()
        root_size = (screen_size[0] / 2, screen_size[1] / 2, screen_size[0] / 4, screen_size[1] / 14)
        root_size = [int(each) for each in root_size]
        self.geometry(f'{root_size[0]}x{root_size[1]}+{root_size[2]}+{root_size[3]}')

        self.frame = ttk.Frame(self)
        self.frame.grid(row=0, column=0, sticky='nsew')
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        def generate_sheet(*args):
            global entry_dict
            global stock_dict

            file = filedialog.asksaveasfilename()
            workbook = Workbook()
            filename = file if file.endswith('xlsx') or file.endswith('xls') else file + '.xlsx'
            workbook.save(filename)

            stock_dict = {}
            for idx, val in enumerate(entry_dict.values()):
                stock_dict[f'Stock {idx + 1}'] = []
                for item in val:
                    stock_dict[f'Stock {idx + 1}'].append(item.get())

            if sheet_name.get() == '':
                edit_spreadsheet(filename, 'Sheet', stock_dict)
            else:
                edit_spreadsheet(filename, sheet_name.get(), stock_dict)

        def generate_entries(*args):
            self.event_generate('<<Generate Entries>>')
            global entry_dict
            entry_dict = {}
            for i in range(stock_num_var.get()):
                ticker_row = 4 + (i * 2)
                ticker_label = ttk.Label(self.frame, text=f"Enter Stock {i + 1}'s Ticker: ")
                ticker_label.grid(row=ticker_row, column=0, padx='5', pady='5', sticky='w')
                ticker_symbol = StringVar()
                entry_dict[f'Stock {i + 1}'] = [ttk.Entry(self.frame, width='6', textvariable=ticker_symbol)]
                entry_dict[f'Stock {i + 1}'][0].grid(row=ticker_row, column=1, pady='5', sticky='w')

                company_row = 5 + (i * 2)
                company_label = ttk.Label(self.frame, text=f"Enter Stock {i + 1}'s Company Name: ")
                company_label.grid(row=company_row, column=0, padx='5', pady='5', sticky='w')
                company_name = StringVar()
                entry_dict[f'Stock {i + 1}'].append(ttk.Entry(self.frame, width='25', textvariable=company_name))
                entry_dict[f'Stock {i + 1}'][1].grid(row=company_row, column=1, sticky='w')

            self.test_button = ttk.Button(self.frame, text='Generate Sheet', command=generate_sheet)
            self.test_button.grid(row=5 + (stock_num_var.get() * 2), column=2, padx='5', pady='5')

        self.button = ttk.Button(self.frame, text='Generate Stock Inputs', command=generate_entries)
        self.button.grid(row=1, column=2, padx='5', pady='5')

        error_var = StringVar()
        error_message = 'Enter a number (ex. 1, 2, 3, 4)'

        def check_numbers(newval, op):
            error_var.set('')
            valid = newval.isdigit() or newval == ''
            if op == 'key':
                valid_so_far = valid and len(newval) <= 2
                if not valid:
                    error_var.set(error_message)
                if not len(newval) <= 2:
                    error_var.set('The number must be less than 100')
                return valid_so_far
            elif op == 'focusout':
                if not valid:
                    error_var.set(error_message)
            return valid

        register = (self.frame.register(check_numbers), '%P', '%V')
        stock_num_var = IntVar()
        self.stock_num = ttk.Entry(self.frame, textvariable=stock_num_var)
        self.stock_num.configure(validate='all', validatecommand=register)
        self.stock_num.grid(row=2, column=1, pady='5', sticky='we')
        self.stock_num_text = ttk.Label(self.frame, text='Amount of Stocks to Add:')
        self.stock_num_text.grid(row=2, column=0, padx='5', sticky='w')

        self.sheet_name = ttk.Label(self.frame, text='Enter Sheet Name: ')
        self.sheet_name.grid(row=1, column=0, padx='5', pady='5', sticky='w')
        sheet_name = StringVar()
        self.sheet_name_entry = ttk.Entry(self.frame, width='20', textvariable=sheet_name)
        self.sheet_name_entry.grid(row=1, column=1, sticky='w')

        self.error = ttk.Label(self.frame, font='TkSmallCaptionFont', foreground='red', textvariable=error_var)
        self.error.grid(row=3, column=1, sticky='w')


class ExistingSpreadsheet(NewSpreadsheet):
    def __init__(self, parent):
        Toplevel.__init__(self, parent)
        self.parent = parent
        self.parent.withdraw()
        self.title('Edit Investment Spreadsheet')

        screen_size = self.wm_maxsize()
        root_size = (screen_size[0] / 2, screen_size[1] / 2, screen_size[0] / 4, screen_size[1] / 14)
        root_size = [int(each) for each in root_size]
        self.geometry(f'{root_size[0]}x{root_size[1]}+{root_size[2]}+{root_size[3]}')

        self.frame = ttk.Frame(self)
        self.frame.grid(row=0, column=0, sticky='nsew')
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        def select_file(*args):
            self.frame.event_generate('<<File Name>>', when='tail')
            file_name = StringVar()
            file_name.set(filedialog.askopenfilename())
            self.frame.bind('<<File Name>>', generate_radiobuttons)
            return file_name

        selected_sheet = StringVar()

        def generate_radiobuttons(*args):
            workbook = load_workbook(filename=file_name.get())
            sheet_names = workbook.sheetnames
            print(sheet_names)
            selected_sheet.set(sheet_names[0])
            for idx, sheet in enumerate(sheet_names):
                ttk.Radiobutton(self.frame, text=sheet, variable=selected_sheet, value=sheet,
                                command=stock_inputs).grid(row=3 + idx, column=0, padx='15', pady='5', sticky='we')
            self.new_sheet = ttk.Radiobutton(self.frame, text='New Sheet', variable=selected_sheet, value='New Sheet',
                                        command=stock_inputs).grid(row=4 + idx, column=0, padx='15', pady='5', sticky='we')

        file_name = select_file()

        new_sheet_var = StringVar()
        stock_amt = IntVar()

        def input_generator(row=1, flag=True):
            if flag:
                try:
                    self.new_sheet_label.destroy()
                    self.new_sheet_entry.destroy()
                    self.stock_amt_text.destroy()
                    self.stock_amt.destroy()
                    self.generate_inputs.destroy()

                    self.new_sheet_label = ttk.Label(self.frame, text='New Sheet Name: ')
                    self.new_sheet_label.grid(row=row, column=1, padx='2', pady='5', sticky='w')
                    self.new_sheet_entry = ttk.Entry(self.frame, width='20', textvariable=new_sheet_var)
                    self.new_sheet_entry.grid(row=row, column=2, pady='5', sticky='w')

                    self.stock_amt_text = ttk.Label(self.frame, text='Enter # of Stocks to Add: ')
                    self.stock_amt_text.grid(row=row + 1, column=1, padx='2', pady='5', sticky='w')

                    self.stock_amt = ttk.Entry(self.frame, textvariable=stock_amt, width='6')
                    self.stock_amt.grid(row=row + 1, column=2, pady='5', sticky='w')

                    self.generate_inputs = ttk.Button(self.frame, text='Generate Stock Inputs',
                                                      command=generate_entries)
                    self.generate_inputs.grid(row=row + 1, column=3, padx='5', pady='5', sticky='w')
                except AttributeError:
                    self.new_sheet_label = ttk.Label(self.frame, text='New Sheet Name: ')
                    self.new_sheet_label.grid(row=row, column=1, padx='2', pady='5', sticky='w')
                    self.new_sheet_entry = ttk.Entry(self.frame, width='20', textvariable=new_sheet_var)
                    self.new_sheet_entry.grid(row=row, column=2, pady='5', sticky='w')

                    self.stock_amt_text = ttk.Label(self.frame, text='Enter # of Stocks to Add: ')
                    self.stock_amt_text.grid(row=row + 1, column=1, padx='2', pady='5', sticky='w')

                    self.stock_amt = ttk.Entry(self.frame, textvariable=stock_amt, width='6')
                    self.stock_amt.grid(row=row + 1, column=2, pady='5', sticky='w')

                    self.generate_inputs = ttk.Button(self.frame, text='Generate Stock Inputs',
                                                      command=generate_entries)
                    self.generate_inputs.grid(row=row + 1, column=3, padx='5', pady='5', sticky='w')

            if not flag:
                try:
                    self.new_sheet_label.destroy()
                    self.new_sheet_entry.destroy()
                    self.stock_amt_text.destroy()
                    self.stock_amt.destroy()
                    self.generate_inputs.destroy()

                    self.stock_amt_text = ttk.Label(self.frame, text='Enter # of Stocks to Add: ')
                    self.stock_amt_text.grid(row=row, column=1, padx='2', pady='5', sticky='w')
                    self.stock_amt = ttk.Entry(self.frame, textvariable=stock_amt, width='6')
                    self.stock_amt.grid(row=row, column=2, pady='5', sticky='w')

                    self.generate_inputs = ttk.Button(self.frame, text='Generate Stock Inputs', command=lambda:generate_entries(flag=False))
                    self.generate_inputs.grid(row=row, column=3, padx='5', pady='5', sticky='w')
                except AttributeError:
                    self.stock_amt_text = ttk.Label(self.frame, text='Enter # of Stocks to Add: ')
                    self.stock_amt_text.grid(row=row, column=1, padx='2', pady='5', sticky='w')
                    self.stock_amt = ttk.Entry(self.frame, textvariable=stock_amt, width='6')
                    self.stock_amt.grid(row=row, column=2, pady='5', sticky='w')

                    self.generate_inputs = ttk.Button(self.frame, text='Generate Stock Inputs',
                                                      command=generate_entries)
                    self.generate_inputs.grid(row=row, column=3, padx='5', pady='5', sticky='w')

        def stock_inputs(*args):
            if selected_sheet.get() == 'New Sheet':
                input_generator()
            else:
                input_generator(flag=False)

        def generate_sheet(*args):
            global entry_dict
            global stock_dict
            print(entry_dict.values())
            stock_dict = {}
            for idx, val in enumerate(entry_dict.values()):
                stock_dict[f'Stock {idx + 1}'] = []
                for item in val:
                    stock_dict[f'Stock {idx + 1}'].append(item.get())

            if selected_sheet.get() == 'New Sheet':
                if new_sheet_var.get() == '':
                    edit_spreadsheet(file_name.get(), '', stock_dict, version=3)
                else:
                    edit_spreadsheet(file_name.get(), new_sheet_var.get(), stock_dict, version=3)
            else:
                edit_spreadsheet(file_name.get(), selected_sheet.get(), stock_dict, version=2)

        def generate_entries(flag=True):
            global entry_dict
            entry_dict = {}
            if flag:
                for i in range(stock_amt.get()):
                    ticker_row = 3 + (i * 2)
                    ticker_label = ttk.Label(self.frame, text=f"Enter Stock {i + 1}'s Ticker: ")
                    ticker_label.grid(row=ticker_row, column=1, padx='5', pady='5', sticky='w')
                    ticker_symbol = StringVar()
                    entry_dict[f'Stock {i + 1}'] = [ttk.Entry(self.frame, width='6', textvariable=ticker_symbol)]
                    entry_dict[f'Stock {i + 1}'][0].grid(row=ticker_row, column=2, pady='5', sticky='w')

                    company_row = 4 + (i * 2)
                    company_label = ttk.Label(self.frame, text=f"Enter Stock {i + 1}'s Company Name: ")
                    company_label.grid(row=company_row, column=1, padx='5', pady='5', sticky='w')
                    company_name = StringVar()
                    entry_dict[f'Stock {i + 1}'].append(ttk.Entry(self.frame, width='20', textvariable=company_name))
                    entry_dict[f'Stock {i + 1}'][1].grid(row=company_row, column=2, sticky='w')

                self.test_button = ttk.Button(self.frame, text='Edit Sheet', command=generate_sheet)
                self.test_button.grid(row=4 + (stock_amt.get() * 2), column=2, padx='5', pady='5')

            if not flag:
                for i in range(stock_amt.get()):
                    ticker_row = 2 + (i * 2)
                    ticker_label = ttk.Label(self.frame, text=f"Enter Stock {i + 1}'s Ticker: ")
                    ticker_label.grid(row=ticker_row, column=1, padx='5', pady='5', sticky='w')
                    ticker_symbol = StringVar()
                    entry_dict[f'Stock {i + 1}'] = [ttk.Entry(self.frame, width='6', textvariable=ticker_symbol)]
                    entry_dict[f'Stock {i + 1}'][0].grid(row=ticker_row, column=2, pady='5', sticky='w')

                    company_row = 3 + (i * 2)
                    company_label = ttk.Label(self.frame, text=f"Enter Stock {i + 1}'s Company Name: ")
                    company_label.grid(row=company_row, column=1, padx='5', pady='5', sticky='w')
                    company_name = StringVar()
                    entry_dict[f'Stock {i + 1}'].append(ttk.Entry(self.frame, width='20', textvariable=company_name))
                    entry_dict[f'Stock {i + 1}'][1].grid(row=company_row, column=2, sticky='w')

                self.test_button = ttk.Button(self.frame, text='Edit Sheet', command=generate_sheet)
                self.test_button.grid(row=3 + (stock_amt.get() * 2), column=2, padx='5', pady='5')

        self.header = ttk.Label(self.frame, text=f'Editing Spreadsheet: {os.path.basename(file_name.get())}')
        self.header.grid(row=0, column=0, columnspan=3, sticky='we')
        self.header.configure(padding='15', font='TkCaptionFont')

        self.sheet_header = ttk.Label(self.frame, text='Sheet Names')
        self.sheet_header.configure(font='TkCaptionFont 13 italic bold')
        self.sheet_header.grid(row=1, column=0, padx='5', pady='5')


if __name__ == '__main__':
    root = Tk()
    MainWindow(root)
    root.mainloop()
