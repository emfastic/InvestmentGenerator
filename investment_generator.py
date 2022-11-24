from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from create_stock import create_stock


def sheet_selection(workbook, sheet_name, version=1):
    """ Asks the user for what sheet to edit if not creating a new one, returns the sheet and sheet_name """
    sheet = None
    if version == 1:
        workbook['Sheet'].title = f'{sheet_name}'
        sheet = workbook[f'{sheet_name}']
        return sheet, sheet_name
    if version == 2:
        sheet = workbook[f'{sheet_name}']
        return sheet, sheet_name
    if version == 3:
        if sheet_name == '':
            sheet_name = workbook.create_sheet().title
            sheet = workbook[f'{sheet_name}']
        else:
            workbook.create_sheet(f'{sheet_name}')
            sheet = workbook[f'{sheet_name}']
        return sheet, sheet_name


def stock_number(sheet_name):
    """ Asks the user for how many stocks they would like to add """
    while True:
        print(f'Enter how many stocks you would like to add to the sheet {sheet_name} below')
        added_stocks = input('> ')
        if added_stocks.isnumeric() and int(added_stocks) > 0:
            added_stocks = int(added_stocks)
            return added_stocks
        print("Make sure you input an integer such as '1', '2', '3', etc.")


def create_stock_list(stock_amt):
    """ Creates a list of tuples in the format (ticker, company name) """
    stock_list = []
    for i in range(stock_amt):
        print(f"Enter stock number {i + 1}'s ticker symbol below.")
        ticker = input('> ').upper()
        print(f"Enter {ticker}'s full company name below.")
        company_name = input('> ')
        stock_list.append((ticker, company_name))
    return stock_list


def create_stock_obj_list(stock_dict):
    """ Creates a list of stock objects """
    stock_obj_list = []
    for stock in stock_dict.values():
        new_stock = create_stock(*stock)
        stock_obj_list.append(new_stock)
    return stock_obj_list


def years_back(n=4):
    """ Ensures that the default number of years (4) of data used is sufficient for the user """
    while True:
        print(f'The default number of years of historical data used is {n}.')
        response = input('Would you like to change how many years back the data goes (y/n): ')
        if response.strip().lower() == 'y':
            historic_range = input('Enter the number of years back: ')
            if historic_range.isnumeric() and int(historic_range) < 15:
                return int(historic_range)
        if response.strip().lower() == 'n':
            return n
        print("Please input y or n. If y make sure to enter a number in the form '1', '2', '3', etc.")


def current_row(sheet):
    """ Returns int val of the row that the new stock will be put on """
    if len(sheet.calculate_dimension()) > 5:
        row_num = sheet.calculate_dimension()[len(sheet.calculate_dimension()) - 2: len(sheet.calculate_dimension())]
    else:
        row_num = sheet.calculate_dimension()[len(sheet.calculate_dimension()) - 1: len(sheet.calculate_dimension())]
    return int(row_num)


def create_header_list(obj_list, years):
    """ Creates list of header names PRECONDITION there is 1 stock in the list """
    header_list = []
    for k, v in vars(obj_list[0]).items():
        if isinstance(v, dict):
            for i in range(years + 1):
                header_list.append(f"{k.upper().replace('_',' ')} {max(v.keys()) - i}")
        else:
            header_list.append(k.upper().replace('_', ' '))
    return header_list


def create_headers(sheet, header_list):
    """ Uses the header list to iterate through each cell in first row and create headers """
    sheet.insert_cols(idx=1, amount=len(header_list) - 1)
    for index, cell in enumerate(sheet[1]):
        cell.value = header_list[index]


def num_formatting(val):
    """ Returns a properly formatted cell '() around neg num' based on whether number is pos or neg or str """
    if isinstance(val, float) and val >= 0 or isinstance(val, str):
        return val
    if isinstance(val, float) and val < 0:
        return f'({abs(val)})'


def range_formatting(y_list):
    """ Returns properly formatted range in form 'min(num) - max(num)' """
    if len(y_list) == 2:
        cell = f'{num_formatting(y_list[0])} - {num_formatting(y_list[1])}'
    else:
        cell = f'{num_formatting(y_list[0])}'
    return cell


def create_stock_val_list(stock, years):
    """ Creates a list of stock values with desired number of years of historical data """
    val_list = list(vars(stock).values())
    stock_val_list = []
    for each in val_list:
        if isinstance(each, dict):
            for i in range(years + 1):
                current_year = max(each.keys())
                stock_val_list.append(each[current_year - i])
        else:
            stock_val_list.append(each)
    return stock_val_list


def create_stock_rows(sheet, stock_list, row_num, years):
    """ Creates a new row with stock data for the amount of stocks needed to be added """
    for stock in stock_list:
        val_list = create_stock_val_list(stock, years)
        for idx, cell in enumerate(sheet[row_num]):
            a_val = val_list[idx]
            if isinstance(a_val, list):
                cell.value = range_formatting(a_val)
            else:
                cell.value = num_formatting(a_val)
        row_num += 1


def edit_spreadsheet(workbook_name, sheet_name, stock_dict, version=1):
    """ Creates investment.xlsx spreadsheet and formats in the desired data """
    workbook = load_workbook(f'{workbook_name}')
    sheet = None
    if version == 1:
        sheet, sheet_name = sheet_selection(workbook, sheet_name)
    if version == 2:
        sheet, sheet_name = sheet_selection(workbook, sheet_name, version)
    if version == 3:
        sheet, sheet_name = sheet_selection(workbook, sheet_name, version)

    stock_obj_list = create_stock_obj_list(stock_dict)
    row_num = current_row(sheet) + 1 # to start on the first blank row
    header_list = create_header_list(stock_obj_list, 4)

    if row_num == 2:
        sheet.insert_rows(idx=1, amount=len(stock_obj_list) + 1)
        create_headers(sheet, header_list)
        create_stock_rows(sheet, stock_obj_list, row_num, 4)
    else:
        sheet.insert_rows(idx=row_num + 1, amount=len(stock_obj_list))
        create_stock_rows(sheet, stock_obj_list, row_num, 4)

    for idx, col in enumerate(sheet.columns, 1):
        sheet.column_dimensions[get_column_letter(idx)].auto_size = True

    workbook.save(f'{workbook_name}')

    return


